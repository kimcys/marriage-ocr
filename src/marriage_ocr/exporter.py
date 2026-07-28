from __future__ import annotations

import csv
from dataclasses import dataclass, replace
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from marriage_ocr.models import ExtractedRecord


XLSX_COLUMNS = [
    "Bil",
    "Nama Suami",
    "IC Lama Suami",
    "IC Baru Suami",
    "ID Suami Raw",
    "Umur Suami",
    "Nama Isteri",
    "IC Lama Isteri",
    "IC Baru Isteri",
    "ID Isteri Raw",
    "Umur Isteri",
    "Mas Kahwin",
    "Mas Kahwin Raw",
    "Nama Pendaftar",
    "Alamat Pendaftar",
    "Nama Wali",
    "Hubungan Wali",
    "Saksi 1",
    "Saksi 2",
    "Tarikh Nikah",
    "Tarikh Nikah Raw",
    "Tarikh Keluar",
    "Tarikh Keluar Raw",
    "Remarks",
    "Confidence",
    "Status Review",
    "Review Reason",
    "Source File",
    "Source Page",
    "Source Record",
    "Crop Folder",
    "Raw Bil",
    "Raw Suami Isteri",
    "Raw Pendaftar",
    "Raw Wali",
    "Raw Hubungan Wali",
    "Raw Saksi",
    "Raw Tarikh Nikah",
    "Raw Tarikh Keluar",
    "Raw Remarks",
    "Raw OCR JSON",
    "Created At",
    "Updated At",
]

PUBLIC_XLSX_COLUMNS = [
    "Bil",
    "Nama Suami",
    "IC Lama Suami",
    "IC Baru Suami",
    "Umur Suami",
    "Nama Isteri",
    "IC Lama Isteri",
    "IC Baru Isteri",
    "Umur Isteri",
    "Mas Kahwin",
    "Nama Pendaftar",
    "Alamat Pendaftar",
    "Nama Wali",
    "Hubungan Wali",
    "Saksi 1",
    "Saksi 2",
    "Tarikh Nikah",
    "Tarikh Keluar",
    "Remarks",
    "Confidence",
    "Status",
    "Created At",
    "Updated At",
]

PUBLIC_CSV_COLUMNS = [column for column in PUBLIC_XLSX_COLUMNS if column not in {"Created At", "Updated At"}]

EXPORT_COLUMN_TO_FIELD = {
    "Bil": "bil",
    "Nama Suami": "nama_suami",
    "IC Lama Suami": "ic_lama_suami",
    "IC Baru Suami": "ic_baru_suami",
    "ID Suami Raw": "id_suami_raw",
    "Umur Suami": "umur_suami",
    "Nama Isteri": "nama_isteri",
    "IC Lama Isteri": "ic_lama_isteri",
    "IC Baru Isteri": "ic_baru_isteri",
    "ID Isteri Raw": "id_isteri_raw",
    "Umur Isteri": "umur_isteri",
    "Mas Kahwin": "mas_kahwin",
    "Mas Kahwin Raw": "mas_kahwin_raw",
    "Nama Pendaftar": "nama_pendaftar",
    "Alamat Pendaftar": "alamat_pendaftar",
    "Nama Wali": "nama_wali",
    "Hubungan Wali": "hubungan_wali",
    "Saksi 1": "saksi_1",
    "Saksi 2": "saksi_2",
    "Tarikh Nikah": "tarikh_nikah",
    "Tarikh Nikah Raw": "tarikh_nikah_raw",
    "Tarikh Keluar": "tarikh_keluar",
    "Tarikh Keluar Raw": "tarikh_keluar_raw",
    "Remarks": "remarks",
    "Confidence": "confidence",
    "Status": "status_review",
    "Status Review": "status_review",
    "Review Reason": "review_reason",
    "Source File": "source_file",
    "Source Page": "source_page",
    "Source Record": "source_record",
    "Crop Folder": "crop_folder",
    "Raw Bil": "raw_bil",
    "Raw Suami Isteri": "raw_suami_isteri",
    "Raw Pendaftar": "raw_pendaftar",
    "Raw Wali": "raw_wali",
    "Raw Hubungan Wali": "raw_hubungan_wali",
    "Raw Saksi": "raw_saksi",
    "Raw Tarikh Nikah": "raw_tarikh_nikah",
    "Raw Tarikh Keluar": "raw_tarikh_keluar",
    "Raw Remarks": "raw_remarks",
    "Raw OCR JSON": "raw_ocr_json",
    "Created At": "created_at",
    "Updated At": "updated_at",
}


@dataclass(frozen=True)
class ExportSummary:
    written_count: int
    skipped_duplicates: int
    total_rows: int
    output_path: Path


def export_records_to_xlsx(
    records: Iterable[ExtractedRecord],
    output_path: Path,
    export_config: Mapping[str, Any],
    *,
    reset_output: bool = False,
    skip_existing: bool = False,
) -> ExportSummary:
    append_mode = bool(export_config.get("append", True))
    dedupe_enabled = bool(export_config.get("dedupe", True)) or skip_existing
    sheet_name = str(export_config.get("sheet_name", "Records"))
    visible_columns, hidden_columns = _resolve_xlsx_columns(export_config)
    sheet_columns = [*visible_columns, *hidden_columns]

    record_list = list(records)
    workbook, worksheet = _load_or_create_workbook(
        output_path=output_path,
        reset_output=reset_output,
        append_mode=append_mode,
        sheet_name=sheet_name,
    )
    _ensure_headers(worksheet, sheet_columns, visible_columns)
    _hide_columns(worksheet, visible_columns, hidden_columns)

    existing_keys = _read_existing_keys(worksheet) if dedupe_enabled else set()
    now = _timestamp_now()
    written_count = 0
    skipped_duplicates = 0

    for record in record_list:
        dedupe_key = _record_dedupe_key(record)
        if dedupe_enabled and dedupe_key is not None and dedupe_key in existing_keys:
            skipped_duplicates += 1
            continue

        export_row = record_to_export_dict(record, timestamp=now, columns=sheet_columns)
        worksheet.append([export_row[column] for column in sheet_columns])
        written_count += 1
        if dedupe_key is not None:
            existing_keys.add(dedupe_key)

    workbook.save(output_path)
    return ExportSummary(
        written_count=written_count,
        skipped_duplicates=skipped_duplicates,
        total_rows=max(0, worksheet.max_row - 1),
        output_path=output_path,
    )


def export_records_to_csv(
    records: Iterable[ExtractedRecord],
    output_path: Path,
    export_config: Mapping[str, Any],
    *,
    reset_output: bool = False,
    skip_existing: bool = False,
) -> ExportSummary:
    append_mode = bool(export_config.get("append", True))
    dedupe_enabled = bool(export_config.get("dedupe", True)) or skip_existing

    record_list = list(records)
    existing_rows: list[dict[str, Any]] = []
    existing_keys: set[tuple[str, ...]] = set()

    if reset_output and output_path.exists():
        output_path.unlink()

    if output_path.exists() and append_mode:
        with output_path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                existing_rows.append({column: row.get(column) for column in PUBLIC_CSV_COLUMNS})
                if dedupe_enabled:
                    key = _csv_row_dedupe_key(row)
                    if key is not None:
                        existing_keys.add(key)

    now = _timestamp_now()
    new_rows: list[dict[str, Any]] = []
    written_count = 0
    skipped_duplicates = 0

    for record in record_list:
        export_row = record_to_export_dict(record, timestamp=now, columns=PUBLIC_CSV_COLUMNS)
        dedupe_key = _csv_record_dedupe_key(record)
        if dedupe_enabled and dedupe_key is not None and dedupe_key in existing_keys:
            skipped_duplicates += 1
            continue
        new_rows.append(export_row)
        written_count += 1
        if dedupe_key is not None:
            existing_keys.add(dedupe_key)

    rows_to_write = [*existing_rows, *new_rows] if append_mode and output_path.exists() else new_rows
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=PUBLIC_CSV_COLUMNS)
        writer.writeheader()
        for row in rows_to_write:
            writer.writerow({column: row.get(column) for column in PUBLIC_CSV_COLUMNS})

    return ExportSummary(
        written_count=written_count,
        skipped_duplicates=skipped_duplicates,
        total_rows=len(rows_to_write),
        output_path=output_path,
    )


def _load_or_create_workbook(
    *,
    output_path: Path,
    reset_output: bool,
    append_mode: bool,
    sheet_name: str,
) -> tuple[Workbook, Worksheet]:
    if reset_output and output_path.exists():
        output_path.unlink()

    if output_path.exists() and append_mode:
        workbook = load_workbook(output_path)
        worksheet = workbook[workbook.sheetnames[0]]
        return workbook, worksheet

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    return workbook, worksheet


def _ensure_headers(worksheet: Worksheet, sheet_columns: list[str], visible_columns: list[str]) -> None:
    if worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet["A1"].value is None:
        for index, value in enumerate(sheet_columns, start=1):
            worksheet.cell(row=1, column=index).value = value
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:{_column_letter(len(visible_columns))}1"
        return

    header = [worksheet.cell(row=1, column=index + 1).value for index in range(len(sheet_columns))]
    if header != sheet_columns:
        raise ValueError("Existing XLSX header does not match expected schema")


def _read_existing_keys(worksheet: Worksheet) -> set[tuple[str, ...]]:
    column_names = [worksheet.cell(row=1, column=index + 1).value for index in range(worksheet.max_column)]
    column_indexes = {str(name): index + 1 for index, name in enumerate(column_names) if name is not None}
    keys: set[tuple[str, ...]] = set()

    for row_index in range(2, worksheet.max_row + 1):
        key = _build_existing_row_key(worksheet, row_index, column_indexes)
        if key is not None:
            keys.add(key)

    return keys


def _with_timestamps(record: ExtractedRecord, timestamp: str) -> ExtractedRecord:
    created_at = record.created_at or timestamp
    updated_at = timestamp
    return replace(record, created_at=created_at, updated_at=updated_at)


def record_to_export_dict(
    record: ExtractedRecord,
    *,
    timestamp: str | None = None,
    columns: Sequence[str] | None = None,
) -> dict[str, Any]:
    record_for_export = _with_timestamps(record, timestamp or _timestamp_now())
    row = dict(zip(XLSX_COLUMNS, _record_to_row(record_for_export), strict=True))
    row["Status"] = row["Status Review"]

    if columns is None:
        return row

    selected: dict[str, Any] = {}
    for column in columns:
        if column not in row:
            raise ValueError(f"Unsupported export column: {column}")
        selected[column] = row[column]
    return selected


def record_from_export_dict(data: Mapping[str, Any]) -> ExtractedRecord:
    payload: dict[str, Any] = {}

    for column, field_name in EXPORT_COLUMN_TO_FIELD.items():
        raw_value = data.get(column)
        if field_name == "review_reason":
            payload[field_name] = _parse_review_reasons(raw_value)
            continue
        if field_name in {"umur_suami", "umur_isteri", "source_page"}:
            parsed_int = _parse_optional_int(raw_value)
            if parsed_int is not None:
                payload[field_name] = parsed_int
            continue
        if field_name == "confidence":
            parsed_float = _parse_optional_float(raw_value)
            if parsed_float is not None:
                payload[field_name] = parsed_float
            continue

        normalized = _normalize_optional_text(raw_value)
        if normalized is not None:
            payload[field_name] = normalized

    return ExtractedRecord(**payload)


def _record_to_row(record: ExtractedRecord) -> list[Any]:
    return [
        record.bil,
        record.nama_suami,
        record.ic_lama_suami,
        record.ic_baru_suami,
        record.id_suami_raw,
        record.umur_suami,
        record.nama_isteri,
        record.ic_lama_isteri,
        record.ic_baru_isteri,
        record.id_isteri_raw,
        record.umur_isteri,
        record.mas_kahwin,
        record.mas_kahwin_raw,
        record.nama_pendaftar,
        record.alamat_pendaftar,
        record.nama_wali,
        record.hubungan_wali,
        record.saksi_1,
        record.saksi_2,
        record.tarikh_nikah,
        record.tarikh_nikah_raw,
        record.tarikh_keluar,
        record.tarikh_keluar_raw,
        record.remarks,
        record.confidence,
        record.status_review,
        "; ".join(record.review_reason),
        record.source_file,
        record.source_page,
        record.source_record,
        record.crop_folder,
        record.raw_bil,
        record.raw_suami_isteri,
        record.raw_pendaftar,
        record.raw_wali,
        record.raw_hubungan_wali,
        record.raw_saksi,
        record.raw_tarikh_nikah,
        record.raw_tarikh_keluar,
        record.raw_remarks,
        record.raw_ocr_json,
        record.created_at,
        record.updated_at,
    ]


def _resolve_xlsx_columns(export_config: Mapping[str, Any]) -> tuple[list[str], list[str]]:
    configured_columns = export_config.get("columns")
    include_raw_columns = bool(export_config.get("include_raw_columns", True))

    if configured_columns is not None:
        visible_columns = _normalize_configured_columns(configured_columns)
    elif include_raw_columns:
        visible_columns = list(XLSX_COLUMNS)
    else:
        visible_columns = list(PUBLIC_XLSX_COLUMNS)

    hidden_columns = [column for column in XLSX_COLUMNS if column not in visible_columns]
    return visible_columns, hidden_columns


def _normalize_configured_columns(configured_columns: Any) -> list[str]:
    if isinstance(configured_columns, (str, bytes)) or not isinstance(configured_columns, Iterable):
        raise ValueError("export.columns must be a sequence of column names")

    visible_columns: list[str] = []
    for column in configured_columns:
        column_name = str(column).strip()
        if not column_name:
            continue
        if column_name not in EXPORT_COLUMN_TO_FIELD:
            raise ValueError(f"Unsupported export column: {column_name}")
        if column_name not in visible_columns:
            visible_columns.append(column_name)

    if not visible_columns:
        raise ValueError("export.columns must contain at least one column")

    return visible_columns


def _hide_columns(worksheet: Worksheet, visible_columns: Sequence[str], hidden_columns: Sequence[str]) -> None:
    for index, _column in enumerate(visible_columns, start=1):
        worksheet.column_dimensions[_column_letter(index)].hidden = False

    for index, _column in enumerate(hidden_columns, start=len(visible_columns) + 1):
        worksheet.column_dimensions[_column_letter(index)].hidden = True


def _build_existing_row_key(
    worksheet: Worksheet,
    row_index: int,
    column_indexes: Mapping[str, int],
) -> tuple[str, ...] | None:
    technical_columns = ("Source File", "Source Page", "Source Record", "Crop Folder")
    if all(column in column_indexes for column in technical_columns):
        source_file = worksheet.cell(row=row_index, column=column_indexes["Source File"]).value
        source_page = worksheet.cell(row=row_index, column=column_indexes["Source Page"]).value
        source_record = worksheet.cell(row=row_index, column=column_indexes["Source Record"]).value
        crop_folder = worksheet.cell(row=row_index, column=column_indexes["Crop Folder"]).value
        key = _build_dedupe_key(
            source_file=source_file,
            source_page=source_page,
            source_record=source_record,
            crop_folder=crop_folder,
        )
        if key is not None:
            return key

    row_values = []
    for column_name, column_index in column_indexes.items():
        if column_name == "Created At":
            continue
        if column_name == "Updated At":
            continue
        row_values.append(str(worksheet.cell(row=row_index, column=column_index).value))
    return tuple(row_values) or None


def _record_dedupe_key(record: ExtractedRecord) -> tuple[str, ...] | None:
    return _build_dedupe_key(
        source_file=record.source_file,
        source_page=record.source_page,
        source_record=record.source_record,
        crop_folder=record.crop_folder,
    )


def _csv_row_dedupe_key(row: Mapping[str, Any]) -> tuple[str, ...] | None:
    values = [row.get(column) for column in PUBLIC_CSV_COLUMNS]
    if any(value not in {None, ""} for value in values):
        return tuple("" if value is None else str(value) for value in values)
    return None


def _csv_record_dedupe_key(record: ExtractedRecord) -> tuple[str, ...] | None:
    values = [record_to_export_dict(record, columns=PUBLIC_CSV_COLUMNS).get(column) for column in PUBLIC_CSV_COLUMNS]
    if any(value not in {None, ""} for value in values):
        return tuple("" if value is None else str(value) for value in values)
    return None


def _build_dedupe_key(
    *,
    source_file: Any,
    source_page: Any,
    source_record: Any,
    crop_folder: Any,
) -> tuple[str, ...] | None:
    if source_file and source_page and source_record:
        return ("source", str(source_file), str(source_page), str(source_record))
    if crop_folder and source_record:
        return ("crop", str(crop_folder), str(source_record))
    return None


def _timestamp_now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _normalize_optional_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_optional_int(value: Any) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _parse_optional_float(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_review_reasons(value: Any) -> list[str]:
    if value is None:
        return []
    return [item.strip() for item in str(value).split(";") if item.strip()]


def _column_letter(column_index: int) -> str:
    letters = ""
    index = column_index
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters
