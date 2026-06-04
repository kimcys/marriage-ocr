from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from marriage_ocr import export_from_postgres as module


class DummyCursor:
    def __init__(self, rows: list[dict[str, object]]):
        self._rows = list(rows)
        self.queries: list[str] = []

    def execute(self, query, params=None):
        self.queries.append(query)

    def fetchmany(self, size: int):
        if not self._rows:
            return []

        chunk = self._rows[:size]
        self._rows = self._rows[size:]
        return chunk

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        return False


class DummyConnection:
    def __init__(self, rows: list[dict[str, object]]):
        self._rows = rows

    def cursor(self):
        return DummyCursor(self._rows)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        return False


def test_export_xlsx_parts_writes_split_files(tmp_path: Path, monkeypatch) -> None:
    rows = [_make_row(index) for index in range(1, 4)]
    monkeypatch.setattr(module, "get_connection", lambda: DummyConnection(rows))

    output_dir = tmp_path / "xlsx_parts"
    module.export_xlsx_parts(str(output_dir), rows_per_file=2)

    part_1 = output_dir / "records_part_001.xlsx"
    part_2 = output_dir / "records_part_002.xlsx"

    assert part_1.exists()
    assert part_2.exists()

    sheet_1 = load_workbook(part_1).active
    sheet_2 = load_workbook(part_2).active

    assert sheet_1.max_row == 3
    assert sheet_2.max_row == 2
    assert sheet_1["A2"].value == 1
    assert sheet_2["A2"].value == 3


def test_export_csv_writes_rows(tmp_path: Path, monkeypatch) -> None:
    rows = [_make_row(index) for index in range(1, 4)]
    monkeypatch.setattr(module, "get_connection", lambda: DummyConnection(rows))

    output_path = tmp_path / "records.csv"
    module.export_csv(str(output_path))

    with output_path.open("r", newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        exported_rows = list(reader)

    assert reader.fieldnames == list(module.EXPORT_COLUMNS)
    assert [row["id"] for row in exported_rows] == ["1", "2", "3"]
    assert exported_rows[0]["source_file"] == "sample_1.jpg"


def _make_row(index: int) -> dict[str, object]:
    return {
        "id": index,
        "source_file": f"sample_{index}.jpg",
        "source_page": 1,
        "source_record": index,
        "bil": str(index),
        "nama_suami": f"HUSBAND {index}",
        "ic_baru_suami": f"900101-10-100{index}",
        "nama_isteri": f"WIFE {index}",
        "ic_baru_isteri": f"900101-10-200{index}",
        "tarikh_nikah": "1994-08-27",
        "mas_kahwin": "RM 80.00",
        "wali": "ABDUL RAHMAN",
        "status": "OK",
        "confidence": 0.95,
        "validation_errors": [],
    }
