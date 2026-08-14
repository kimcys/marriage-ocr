import threading
import time
from pathlib import Path
import csv

import pytest
from openpyxl import load_workbook

from marriage_ocr import exporter as exporter_module
from marriage_ocr.exporter import (
    PUBLIC_CSV_COLUMNS,
    PUBLIC_XLSX_COLUMNS,
    XLSX_COLUMNS,
    _atomic_write_bytes,
    export_records_to_csv,
    export_records_to_xlsx,
    record_from_export_dict,
    record_to_export_dict,
)
from marriage_ocr.models import ExtractedRecord


EXPORT_CONFIG = {
    "append": True,
    "dedupe": True,
    "sheet_name": "Records",
}


def test_exporter_writes_schema_and_row(tmp_path: Path) -> None:
    output_path = tmp_path / "records.xlsx"
    record = _make_record(source_record="record_001")

    summary = export_records_to_xlsx([record], output_path, EXPORT_CONFIG, reset_output=True)

    workbook = load_workbook(output_path)
    worksheet = workbook.active

    assert summary.written_count == 1
    assert summary.skipped_duplicates == 0
    assert worksheet.max_row == 2
    assert [worksheet.cell(row=1, column=index + 1).value for index in range(len(XLSX_COLUMNS))] == XLSX_COLUMNS
    assert worksheet["A2"].value == "1"
    assert worksheet["B2"].value == "MOHAMAD BIN YASMIN"


def test_exporter_writes_csv_schema_and_row(tmp_path: Path) -> None:
    output_path = tmp_path / "records.csv"
    record = _make_record(source_record="record_001")

    summary = export_records_to_csv([record], output_path, EXPORT_CONFIG, reset_output=True)

    with output_path.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)

    assert summary.written_count == 1
    assert summary.skipped_duplicates == 0
    assert reader.fieldnames == PUBLIC_CSV_COLUMNS
    assert rows[0]["Bil"] == "1"
    assert rows[0]["Nama Suami"] == "MOHAMAD BIN YASMIN"
    assert "Source Record" not in rows[0]
    assert "Source File" not in rows[0]
    assert "Created At" not in rows[0]
    assert "Updated At" not in rows[0]


def test_exporter_can_show_public_columns_only(tmp_path: Path) -> None:
    output_path = tmp_path / "records.xlsx"
    record = _make_record(source_record="record_001")
    export_config = {
        "append": True,
        "dedupe": True,
        "sheet_name": "Records",
        "include_raw_columns": False,
        "columns": PUBLIC_XLSX_COLUMNS,
    }

    summary = export_records_to_xlsx([record], output_path, export_config, reset_output=True)

    workbook = load_workbook(output_path)
    worksheet = workbook.active

    assert summary.written_count == 1
    assert [worksheet.cell(row=1, column=index + 1).value for index in range(len(PUBLIC_XLSX_COLUMNS))] == PUBLIC_XLSX_COLUMNS
    assert worksheet.column_dimensions["X"].hidden is True
    assert worksheet.column_dimensions["A"].hidden is False
    expected_hidden = [column for column in XLSX_COLUMNS if column not in PUBLIC_XLSX_COLUMNS]
    assert worksheet.max_column == len(PUBLIC_XLSX_COLUMNS) + len(expected_hidden)
    assert worksheet["X1"].value == "ID Suami Raw"


def test_exporter_skips_duplicates_on_rerun(tmp_path: Path) -> None:
    output_path = tmp_path / "records.xlsx"
    record = _make_record(source_record="record_001")

    first = export_records_to_xlsx([record], output_path, EXPORT_CONFIG, reset_output=True)
    second = export_records_to_xlsx([record], output_path, EXPORT_CONFIG, reset_output=False)

    workbook = load_workbook(output_path)
    worksheet = workbook.active

    assert first.written_count == 1
    assert second.written_count == 0
    assert second.skipped_duplicates == 1
    assert worksheet.max_row == 2


def test_concurrent_csv_exports_do_not_lose_rows(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    """Regression: export_records_to_csv used to read-merge-write with no
    locking, so two workers exporting into the same shared output path at
    once could race -- the second worker reads the file before the first
    worker's write lands, then writes back its own merge, silently
    discarding the first worker's newly-appended row. Real scenario: two
    Celery workers finishing pages of the same document and both appending
    to one export file around the same time.

    _timestamp_now is patched to sleep, widening the race window
    deterministically instead of relying on incidental thread timing.
    """
    output_path = tmp_path / "records.csv"
    record_a = _make_record(source_record="record_A", bil="1")
    record_b = _make_record(source_record="record_B", bil="2")

    original_timestamp_now = exporter_module._timestamp_now

    def slow_timestamp_now() -> str:
        time.sleep(0.05)
        return original_timestamp_now()

    monkeypatch.setattr(exporter_module, "_timestamp_now", slow_timestamp_now)

    errors: list[Exception] = []

    def run(record: ExtractedRecord) -> None:
        try:
            export_records_to_csv([record], output_path, EXPORT_CONFIG)
        except Exception as exc:  # pragma: no cover - surfaced via errors list
            errors.append(exc)

    first = threading.Thread(target=run, args=(record_a,))
    first.start()
    time.sleep(0.01)  # let the first worker start its (slow) critical section
    second = threading.Thread(target=run, args=(record_b,))
    second.start()
    first.join()
    second.join()

    assert not errors
    with output_path.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    assert {row["Bil"] for row in rows} == {"1", "2"}


def test_atomic_write_bytes_preserves_original_file_on_failure(tmp_path: Path) -> None:
    """Regression: exporter writes used to go directly to the real output
    path. A crash mid-write (OOM, Celery hard timeout, disk full) truncated
    or corrupted the file, destroying every previously accumulated row, not
    just the new batch. _atomic_write_bytes must leave the original file
    untouched -- and leave no stray temp file behind -- when write_fn fails.
    """
    output_path = tmp_path / "records.csv"
    output_path.write_text("original", encoding="utf-8")

    def failing_write(path: Path) -> None:
        Path(path).write_text("partial", encoding="utf-8")
        raise RuntimeError("simulated crash mid-write")

    with pytest.raises(RuntimeError):
        _atomic_write_bytes(output_path, failing_write)

    assert output_path.read_text(encoding="utf-8") == "original"
    assert list(tmp_path.glob(f".{output_path.name}.*.tmp")) == []


def test_exporter_reset_output_replaces_existing_rows(tmp_path: Path) -> None:
    output_path = tmp_path / "records.xlsx"
    first_record = _make_record(source_record="record_001", bil="1")
    second_record = _make_record(source_record="record_002", bil="2")

    export_records_to_xlsx([first_record], output_path, EXPORT_CONFIG, reset_output=True)
    summary = export_records_to_xlsx([second_record], output_path, EXPORT_CONFIG, reset_output=True)

    workbook = load_workbook(output_path)
    worksheet = workbook.active

    assert summary.written_count == 1
    assert worksheet.max_row == 2
    assert worksheet["A2"].value == "2"


def test_exporter_round_trips_export_dict() -> None:
    record = _make_record(source_record="record_001")

    restored = record_from_export_dict(record_to_export_dict(record, timestamp="2026-05-25T12:00:00"))

    assert restored.bil == record.bil
    assert restored.nama_suami == record.nama_suami
    assert restored.umur_suami == record.umur_suami
    assert restored.confidence == record.confidence
    assert restored.status_review == record.status_review
    assert restored.source_page == record.source_page


def test_exporter_round_trips_public_export_dict() -> None:
    record = _make_record(source_record="record_001")

    restored = record_from_export_dict(
        record_to_export_dict(record, timestamp="2026-05-25T12:00:00", columns=PUBLIC_XLSX_COLUMNS)
    )

    assert restored.bil == record.bil
    assert restored.status_review == record.status_review
    assert restored.created_at == "2026-05-25T12:00:00"


def _make_record(*, source_record: str, bil: str = "1") -> ExtractedRecord:
    return ExtractedRecord(
        bil=bil,
        nama_suami="MOHAMAD BIN YASMIN",
        ic_lama_suami="A.1192345",
        umur_suami=25,
        nama_isteri="SITI BINTI ALI",
        ic_baru_isteri="900101-10-1234",
        umur_isteri=23,
        mas_kahwin="RM 80.00",
        mas_kahwin_raw="RM 8O.OO",
        nama_pendaftar="MOHD SALLEH",
        alamat_pendaftar="KAMPUNG BARU",
        nama_wali="ABDUL RAHMAN",
        hubungan_wali="BAPA",
        saksi_1="AHMAD BIN ALI",
        saksi_2="OSMAN BIN DIN",
        tarikh_nikah="27-08-1994",
        tarikh_nikah_raw="27.8.94",
        tarikh_keluar="02-06-1995",
        tarikh_keluar_raw="2.6.95",
        remarks="TIADA",
        confidence=0.93,
        status_review="OK",
        review_reason=[],
        source_file="sample.jpg",
        source_page=1,
        source_record=source_record,
        crop_folder=f"debug/sample/{source_record}",
        raw_bil=bil,
        raw_suami_isteri="RAW SUAMI ISTERI",
        raw_pendaftar="RAW PENDAFTAR",
        raw_wali="RAW WALI",
        raw_hubungan_wali="RAW HUBUNGAN",
        raw_saksi="RAW SAKSI",
        raw_tarikh_nikah="RAW NIKAH",
        raw_tarikh_keluar="RAW KELUAR",
        raw_remarks="RAW REMARKS",
        raw_ocr_json="{\"mock\":true}",
    )
