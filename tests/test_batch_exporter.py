import csv
from pathlib import Path

from openpyxl import load_workbook

from marriage_ocr.batch_exporter import (
    export_records_to_csv,
    export_records_to_csv_parts,
    export_records_to_xlsx_parts,
)
from marriage_ocr.exporter import XLSX_COLUMNS
from marriage_ocr.models import ExtractedRecord


EXPORT_CONFIG = {
    "append": True,
    "dedupe": True,
    "sheet_name": "Records",
}


def test_export_records_to_csv_writes_expected_columns(tmp_path: Path) -> None:
    output_path = tmp_path / "records.csv"
    record = _make_record(source_record="record_001")

    written_path = export_records_to_csv([record], output_path)

    with written_path.open("r", newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)

    assert written_path == output_path
    assert reader.fieldnames == XLSX_COLUMNS
    assert len(rows) == 1
    assert rows[0]["Bil"] == "1"
    assert rows[0]["Nama Suami"] == "MOHAMAD BIN YASMIN"
    assert rows[0]["Review Reason"] == ""
    assert rows[0]["Created At"] != ""
    assert rows[0]["Updated At"] != ""


def test_export_records_to_xlsx_parts_splits_into_expected_files(tmp_path: Path) -> None:
    records = [_make_record(source_record=f"record_{index:03d}", bil=str(index)) for index in range(1, 6)]

    paths = export_records_to_xlsx_parts(
        records,
        tmp_path / "xlsx_parts",
        EXPORT_CONFIG,
        rows_per_file=2,
    )

    assert [path.name for path in paths] == [
        "records_part_001.xlsx",
        "records_part_002.xlsx",
        "records_part_003.xlsx",
    ]

    first = load_workbook(paths[0]).active
    second = load_workbook(paths[1]).active
    third = load_workbook(paths[2]).active

    assert first.max_row == 3
    assert second.max_row == 3
    assert third.max_row == 2
    assert first["A2"].value == "1"
    assert second["A2"].value == "3"
    assert third["A2"].value == "5"


def test_export_records_to_csv_parts_splits_into_expected_files(tmp_path: Path) -> None:
    records = [_make_record(source_record=f"record_{index:03d}", bil=str(index)) for index in range(1, 6)]

    paths = export_records_to_csv_parts(
        records,
        tmp_path / "csv_parts",
        rows_per_file=2,
    )

    assert [path.name for path in paths] == [
        "records_part_001.csv",
        "records_part_002.csv",
        "records_part_003.csv",
    ]

    with paths[0].open("r", newline="", encoding="utf-8") as handle:
        first_rows = list(csv.DictReader(handle))
    with paths[1].open("r", newline="", encoding="utf-8") as handle:
        second_rows = list(csv.DictReader(handle))
    with paths[2].open("r", newline="", encoding="utf-8") as handle:
        third_rows = list(csv.DictReader(handle))

    assert [row["Bil"] for row in first_rows] == ["1", "2"]
    assert [row["Bil"] for row in second_rows] == ["3", "4"]
    assert [row["Bil"] for row in third_rows] == ["5"]


def test_batch_exporter_handles_empty_records(tmp_path: Path) -> None:
    csv_path = tmp_path / "records.csv"
    written_csv = export_records_to_csv([], csv_path)
    xlsx_paths = export_records_to_xlsx_parts([], tmp_path / "xlsx_parts", EXPORT_CONFIG, rows_per_file=2)

    with written_csv.open("r", newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)

    assert reader.fieldnames == XLSX_COLUMNS
    assert rows == []
    assert xlsx_paths == []


def _make_record(*, source_record: str, bil: str = "1") -> ExtractedRecord:
    return ExtractedRecord(
        bil=bil,
        nama_suami="MOHAMAD BIN YASMIN",
        ic_lama_suami="A.1192345",
        umur_suami=25,
        nama_isteri="SITI BINTI ALI",
        ic_baru_isteri="900101-10-1234",
        umur_isteri=23,
        mas_kahwin="RM 80.00",
        mas_kahwin_raw="RM 8O.OO",
        nama_pendaftar="MOHD SALLEH",
        alamat_pendaftar="KAMPUNG BARU",
        nama_wali="ABDUL RAHMAN",
        hubungan_wali="BAPA",
        saksi_1="AHMAD BIN ALI",
        saksi_2="OSMAN BIN DIN",
        tarikh_nikah="27-08-1994",
        tarikh_nikah_raw="27.8.94",
        tarikh_keluar="02-06-1995",
        tarikh_keluar_raw="2.6.95",
        remarks="TIADA",
        confidence=0.93,
        status_review="OK",
        review_reason=[],
        source_file="sample.jpg",
        source_page=1,
        source_record=source_record,
        crop_folder=f"debug/sample/{source_record}",
        raw_bil=bil,
        raw_suami_isteri="RAW SUAMI ISTERI",
        raw_pendaftar="RAW PENDAFTAR",
        raw_wali="RAW WALI",
        raw_hubungan_wali="RAW HUBUNGAN",
        raw_saksi="RAW SAKSI",
        raw_tarikh_nikah="RAW NIKAH",
        raw_tarikh_keluar="RAW KELUAR",
        raw_remarks="RAW REMARKS",
        raw_ocr_json='{"mock":true}',
    )
