from pathlib import Path
import json

from openpyxl import load_workbook

from marriage_ocr.models import ExtractedRecord
from marriage_ocr.review_store import (
    discover_record_directories,
    export_reviewed_records,
    load_review_bundle,
    save_corrected_record,
)


EXPORT_CONFIG = {
    "append": False,
    "dedupe": False,
    "sheet_name": "Records",
}


def test_review_store_loads_validated_record_and_cell_paths(tmp_path: Path) -> None:
    record_dir = _create_record_dir(tmp_path, source_record="record_001", bil="1")

    bundle = load_review_bundle(record_dir)

    assert bundle.active_record.bil == "1"
    assert bundle.corrected_record is None
    assert bundle.corrected_cells == {}
    assert bundle.verified is False
    assert bundle.full_record_path == record_dir / "full_record.jpg"
    assert set(bundle.cell_paths) == {"bil", "suami_isteri"}
    assert bundle.active_cell_labels["bil"] == "1"
    assert bundle.active_cell_labels["suami_isteri"] == "RAW"


def test_review_store_saves_and_reloads_corrected_record(tmp_path: Path) -> None:
    record_dir = _create_record_dir(tmp_path, source_record="record_001", bil="1")
    corrected_record = _make_record(source_record="record_001", bil="99", status_review="OK")

    save_corrected_record(
        record_dir,
        corrected_record,
        verified=True,
        reviewed_by="QA User",
        review_notes="Checked against source image",
        corrected_cells={"bil": "99", "suami_isteri": "CORRECTED RAW"},
    )

    bundle = load_review_bundle(record_dir)

    assert bundle.active_record.bil == "99"
    assert bundle.corrected_record is not None
    assert bundle.corrected_cells == {"bil": "99", "suami_isteri": "CORRECTED RAW"}
    assert bundle.active_cell_labels["suami_isteri"] == "CORRECTED RAW"
    assert bundle.verified is True
    assert bundle.reviewed_by == "QA User"
    assert bundle.review_notes == "Checked against source image"


def test_review_export_uses_corrected_records_and_verified_filter(tmp_path: Path) -> None:
    first_record_dir = _create_record_dir(tmp_path, source_record="record_001", bil="1")
    second_record_dir = _create_record_dir(tmp_path, source_record="record_002", bil="2")

    save_corrected_record(
        first_record_dir,
        _make_record(source_record="record_001", bil="100", status_review="OK"),
        verified=True,
        reviewed_by="QA User",
        corrected_cells={"bil": "100", "suami_isteri": "TRAIN LABEL 1"},
    )
    save_corrected_record(
        second_record_dir,
        _make_record(source_record="record_002", bil="200", status_review="REVIEW"),
        verified=False,
        reviewed_by="QA User",
        corrected_cells={"bil": "200", "suami_isteri": "TRAIN LABEL 2"},
    )

    output_path = tmp_path / "reviewed.xlsx"
    summary = export_reviewed_records(
        tmp_path,
        output_path,
        EXPORT_CONFIG,
        verified_only=True,
        reset_output=True,
    )

    workbook = load_workbook(output_path)
    worksheet = workbook.active

    assert summary.written_count == 1
    assert worksheet.max_row == 2
    assert worksheet["A2"].value == "100"


def test_discover_record_directories_returns_sorted_records(tmp_path: Path) -> None:
    later = _create_record_dir(tmp_path, source_record="record_010", bil="10", page_name="page_b")
    earlier = _create_record_dir(tmp_path, source_record="record_002", bil="2", page_name="page_a")

    records = discover_record_directories(tmp_path)

    assert records == [earlier, later]


def _create_record_dir(
    root: Path,
    *,
    source_record: str,
    bil: str,
    page_name: str = "page_a",
) -> Path:
    record_dir = root / page_name / "records" / source_record
    record_dir.mkdir(parents=True)

    validated_record = _make_record(source_record=source_record, bil=bil)
    (record_dir / "validated_record.json").write_text(
        json.dumps(validated_record.to_dict(), indent=2, ensure_ascii=True),
        encoding="utf-8",
    )
    (record_dir / "parsed_record.json").write_text(
        json.dumps(validated_record.to_dict(), indent=2, ensure_ascii=True),
        encoding="utf-8",
    )
    (record_dir / "raw_ocr.json").write_text(
        json.dumps({"cells": {"bil": {"text": bil}, "suami_isteri": {"text": "RAW"}}}, indent=2, ensure_ascii=True),
        encoding="utf-8",
    )
    (record_dir / "full_record.jpg").write_bytes(b"full")
    (record_dir / "bil.jpg").write_bytes(b"bil")
    (record_dir / "suami_isteri.jpg").write_bytes(b"suami_isteri")
    return record_dir


def _make_record(*, source_record: str, bil: str, status_review: str = "REVIEW") -> ExtractedRecord:
    return ExtractedRecord(
        bil=bil,
        nama_suami="MOHAMAD BIN YASMIN",
        ic_lama_suami="A.1192345",
        umur_suami=25,
        nama_isteri="SITI BINTI ALI",
        ic_baru_isteri="900101-10-1234",
        umur_isteri=23,
        mas_kahwin="RM 80.00",
        mas_kahwin_raw="RM 80.00",
        nama_pendaftar="MOHD SALLEH",
        alamat_pendaftar="KAMPUNG BARU",
        nama_wali="ABDUL RAHMAN",
        hubungan_wali="BAPA",
        saksi_1="AHMAD BIN ALI",
        saksi_2="OSMAN BIN DIN",
        tarikh_nikah="27-08-1994",
        tarikh_nikah_raw="27.8.94",
        tarikh_keluar="02-06-1995",
        tarikh_keluar_raw="2.6.95",
        remarks="TIADA",
        confidence=0.93,
        status_review=status_review,
        review_reason=[],
        source_file="sample.jpg",
        source_page=1,
        source_record=source_record,
        crop_folder=f"debug/sample/{source_record}",
        raw_bil=bil,
        raw_suami_isteri="RAW SUAMI ISTERI",
        raw_pendaftar="RAW PENDAFTAR",
        raw_wali="RAW WALI",
        raw_hubungan_wali="RAW HUBUNGAN",
        raw_saksi="RAW SAKSI",
        raw_tarikh_nikah="RAW NIKAH",
        raw_tarikh_keluar="RAW KELUAR",
        raw_remarks="RAW REMARKS",
        raw_ocr_json="{\"mock\":true}",
    )
